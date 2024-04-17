"""
    Program to find duplicates in columns
    for a given excel sheet and creates a new sheet in the same excel file
    to write duplicates
"""
import pandas as pd
import os
import openpyxl


def compare_columns(file):
    if check_file_exists(file):

        # Read excel file using pandas and convert to DataFrame
        df = pd.read_excel(file, header=None)

        # df.columns returns integers.
        # Adding the integer value from df.columns to ASCII code to identify the Column Name - (Ex - A, B, C, D etc.)
        default_headers = list(chr(ord('A') + col_number) for col_number in range(len(df.columns)))

        # Convert dataframe to a sorted list
        sorted_list = sorted(df.iloc[:, :len(default_headers)].astype(str).values.flatten())

        # filter string with 'nan' pd.read_excel() loads the file to dataFrame Object. If there are any empty cells
        # in columns and rows. Then, DataFrame sets as Nan. From line 22, we convert dataframe to list of type String.
        # This sets Nan to 'Nan'
        # Lambda expression to create a new clean_list and exclude string 'nan'
        clean_list = list(filter(lambda item: item != 'nan', sorted_list))

        # find duplicates from a list
        dup_list = find_duplicates(clean_list)

        # to write duplicates to the current excel sheet with new sheet 'Duplicates' and add duplicates values
        write_duplicates(dup_list, file)

    else:
        raise Exception('File does not exists')


def find_duplicates(final_list):
    """
    Method to identify the duplicates and creates a set to return
    :param final_list:
    :return: set
    """
    seen = set()
    duplicates = set()

    for item in final_list:
        if item in seen:
            duplicates.add(item)
        else:
            seen.add(item)

    return duplicates


def write_duplicates(dup_set, file):
    workbook = openpyxl.load_workbook(file)
    duplicate_sheet = workbook.create_sheet('Duplicates')
    workbook.active = duplicate_sheet
    count = 1

    for item in dup_set:
        duplicate_sheet.cell(row=count, column=1, value=str(item))
        count += 1

    try:
        workbook.save(file)

    except PermissionError:
        print(f'Oops file is open on your machine. Please close the file and try again')
    except Exception as e:
        print(f'Error Occurred - ', {e})


def check_file_exists(file) -> bool:
    return os.path.isfile(file)


def main():
    compare_columns('applist - Copy.xlsx')


if __name__ == '__main__':
    main()
